"""
Manual Tracking Module - Dartfish-like tracking system
"""
import csv
from datetime import datetime
from typing import List, Dict, Optional
from dataclasses import dataclass, asdict
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Import translation system
try:
    from frontend.translations import get_translation_manager, t
    TRANSLATIONS_AVAILABLE = True
except ImportError:
    TRANSLATIONS_AVAILABLE = False
    def t(text, default=None):
        return default if default else text

@dataclass
class TrackingEvent:
    """Data class for tracking events"""
    timestamp: float  # Time in seconds
    event_type: str  # Goal, Shot, Pass, Tackle, Foul, etc.
    team: str  # Team A, Team B, or Neutral
    outcome: Optional[str] = None  # Result of the action (สำเร็จ, ไม่สำเร็จ, etc.)
    player_number: Optional[int] = None
    player_name: Optional[str] = None  # ชื่อนักเตะ (สำหรับประตู)
    description: str = ""
    half: int = 1  # 1 or 2
    x_position: Optional[float] = None
    y_position: Optional[float] = None
    
    def to_dict(self) -> Dict:
        return asdict(self)

class ManualTrackingData:
    """Manages manual tracking data"""
    def __init__(self):
        self.events: List[TrackingEvent] = []
        self.video_path: Optional[str] = None
        self.match_name: str = "Match"
        self.team_a_name: str = "Team A"
        self.team_b_name: str = "Team B"
        self.team_a_players: List[str] = []  # รายชื่อนักเตะทีม 1
        self.team_b_players: List[str] = []  # รายชื่อนักเตะทีม 2
    
    def _t(self, text: str, default: Optional[str] = None) -> str:
        """Get translated text based on current language"""
        if TRANSLATIONS_AVAILABLE:
            translation_manager = get_translation_manager()
            # Use translate method directly to ensure correct language
            return translation_manager.translate(text, default or text)
        return default or text
    
    def _get_current_language(self) -> str:
        """Get current language setting"""
        if TRANSLATIONS_AVAILABLE:
            translation_manager = get_translation_manager()
            return translation_manager.get_language()
        return "TH"  # Default to Thai
    
    def _get_th_text(self, th_text: str, en_text: str) -> str:
        """Get Thai text if current language is TH, otherwise English"""
        current_lang = self._get_current_language()
        return th_text if current_lang == "TH" else en_text
    
    def _get_current_language(self) -> str:
        """Get current language setting"""
        if TRANSLATIONS_AVAILABLE:
            translation_manager = get_translation_manager()
            return translation_manager.get_language()
        return "TH"  # Default to Thai
    
    def _get_th_text(self, th_text: str, en_text: str) -> str:
        """Get Thai text if current language is TH, otherwise English"""
        current_lang = self._get_current_language()
        return th_text if current_lang == "TH" else en_text
    
    def add_event(self, event: TrackingEvent):
        """Add a tracking event"""
        self.events.append(event)
        # Sort by timestamp
        self.events.sort(key=lambda x: x.timestamp)
    
    def remove_event(self, index: int):
        """Remove event by index"""
        if 0 <= index < len(self.events):
            self.events.pop(index)
    
    def get_events_by_type(self, event_type: str) -> List[TrackingEvent]:
        """Get all events of a specific type"""
        return [e for e in self.events if e.event_type == event_type]
    
    def get_events_by_team(self, team: str) -> List[TrackingEvent]:
        """Get all events for a specific team"""
        return [e for e in self.events if e.team == team]
    
    def export_to_csv(self, filepath: str):
        """Export tracking data to CSV (legacy method)"""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            if not self.events:
                return
            
            writer = csv.DictWriter(f, fieldnames=self.events[0].to_dict().keys())
            writer.writeheader()
            for event in self.events:
                writer.writerow(event.to_dict())
    
    def export_to_excel(self, filepath: str):
        """Export tracking data to Excel with multiple sheets"""
        if not PANDAS_AVAILABLE:
            # Fallback to CSV if pandas not available
            self.export_to_csv(filepath.replace('.xlsx', '.csv'))
            return
        
        if not self.events:
            return
        
        # Get current language to ensure consistent translation
        current_lang = self._get_current_language()
        
        # Define column names based on language
        if current_lang == "EN":
            time_col = "Time"
            time_sec_col = "Time (seconds)"
            half_col = "Half"
            event_col = "Event"
            outcome_col = "Outcome"
            team_col = "Team"
            half_text_col = "Half (text)"
            half_num_col = "Half (number)"
            player_num_col = "Player Number"
            player_name_col = "Player Name"
            description_col = "Description"
            x_pos_col = "X Position"
            y_pos_col = "Y Position"
            
            half_mapping = {
                1: "First Half",
                2: "Second Half",
                3: "Extra Time First Half",
                4: "Extra Time Second Half"
            }
            neutral_text = "Neutral"
        else:  # TH
            time_col = "เวลา"
            time_sec_col = "เวลา (วินาที)"
            half_col = "ครึ่ง"
            event_col = "เหตุการณ์"
            outcome_col = "ผลลัพธ์"
            team_col = "ทีม"
            half_text_col = "ครึ่ง (ข้อความ)"
            half_num_col = "ครึ่ง (ตัวเลข)"
            player_num_col = "หมายเลขผู้เล่น"
            player_name_col = "ชื่อผู้เล่น"
            description_col = "คำอธิบาย"
            x_pos_col = "ตำแหน่ง X"
            y_pos_col = "ตำแหน่ง Y"
            
            half_mapping = {
                1: "ครึ่งแรก",
                2: "ครึ่งหลัง",
                3: "ต่อเวลาครึ่งแรก",
                4: "ต่อเวลาครึ่งหลัง"
            }
            neutral_text = "กลาง"
        
        # Prepare data
        data = []
        for event in self.events:
            event_dict = event.to_dict()
            # Format timestamp
            minutes = int(event.timestamp // 60)
            seconds = int(event.timestamp % 60)
            event_dict[time_col] = f"{minutes:02d}:{seconds:02d}"
            event_dict[time_sec_col] = event.timestamp
            # Map half to text
            event_dict[half_col] = half_mapping.get(event.half, f"{half_col} {event.half}")
            # Ensure player_name is included
            if 'player_name' not in event_dict:
                event_dict['player_name'] = None
            data.append(event_dict)
        
        # Sheet 1: Raw Data - Translate event_type and outcome values
        for item in data:
            if 'event_type' in item and item['event_type']:
                item['event_type'] = self._t(item['event_type'], item['event_type'])
            if 'outcome' in item and item['outcome']:
                item['outcome'] = self._t(item['outcome'], item['outcome'])
            if 'team' in item and item['team']:
                # Translate "กลาง" to "Neutral" if needed
                if item['team'] == "กลาง":
                    item['team'] = neutral_text
        
        df_raw = pd.DataFrame(data)
        # Reorder columns
        column_order = [time_col, time_sec_col, 'event_type', 'outcome', 'team', half_col, 'half', 
                       'player_number', 'player_name', 'description', 'x_position', 'y_position']
        df_raw = df_raw[[col for col in column_order if col in df_raw.columns]]
        df_raw.columns = [
            time_col, 
            time_sec_col, 
            event_col, 
            outcome_col, 
            team_col, 
            half_text_col, 
            half_num_col, 
            player_num_col, 
            player_name_col, 
            description_col, 
            x_pos_col, 
            y_pos_col
        ]
        
        # Get current language to ensure consistent translation
        current_lang = self._get_current_language()
        
        # Sheet 2: Summary (All)
        summary_all = self._generate_summary(self.events)
        
        # Sheet 3: Team Comparison (only if 2 teams exist)
        comparison = self._generate_comparison(self.events)
        
        # Define sheet names based on language
        if current_lang == "EN":
            raw_data_sheet = "Raw Data"
            summary_sheet = "Summary"
            comparison_sheet = "Team Comparison"
            players_sheet = "Players List"
            goals_sheet = "Goals Summary"
            timeline_sheet = "Timeline"
            key_moments_sheet = "Key Moments"
            event_freq_sheet = "Event Frequency"
            set_pieces_sheet = "Set Pieces"
            player_stats_sheet = "Player Statistics"
            
            half_sheet_names = {
                1: "First Half",
                2: "Second Half",
                3: "Extra Time First Half",
                4: "Extra Time Second Half"
            }
        else:  # TH
            raw_data_sheet = "ข้อมูลดิบ"
            summary_sheet = "สรุปผลรวม"
            comparison_sheet = "เปรียบเทียบทีม"
            players_sheet = "รายชื่อนักเตะ"
            goals_sheet = "สรุปประตู"
            timeline_sheet = "ไทม์ไลน์"
            key_moments_sheet = "ช่วงเวลาสำคัญ"
            event_freq_sheet = "ความถี่เหตุการณ์"
            set_pieces_sheet = "ลูกตั้งเตะ"
            player_stats_sheet = "สถิติผู้เล่น"
            
            half_sheet_names = {
                1: "ครึ่งแรก",
                2: "ครึ่งหลัง",
                3: "ต่อเวลาครึ่งแรก",
                4: "ต่อเวลาครึ่งหลัง"
            }
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Sheet 1: Raw Data
            df_raw.to_excel(writer, sheet_name=raw_data_sheet, index=False)
            self._apply_sheet_styling(writer, raw_data_sheet)
            
            # Sheet 2: Summary All
            summary_all.to_excel(writer, sheet_name=summary_sheet, index=False)
            self._apply_sheet_styling(writer, summary_sheet)
            
            # Sheet 3: Team Comparison (only if 2 teams exist)
            if not comparison.empty:
                comparison.to_excel(writer, sheet_name=comparison_sheet, index=False)
                self._apply_sheet_styling(writer, comparison_sheet)
            
            # Sheet: Players List (only if players are provided)
            if self.team_a_players or self.team_b_players:
                players_data = []
                max_players = max(len(self.team_a_players), len(self.team_b_players))
                
                # Use actual team names
                team_a_header = self.team_a_name if self.team_a_name else self._get_th_text("ทีม 1", "Team 1")
                team_b_header = self.team_b_name if self.team_b_name else self._get_th_text("ทีม 2", "Team 2")
                
                order_col = self._get_th_text("ลำดับ", "Order")
                
                for i in range(max_players):
                    team1_player = self.team_a_players[i] if i < len(self.team_a_players) else ''
                    team2_player = self.team_b_players[i] if i < len(self.team_b_players) else ''
                    
                    row = {
                        order_col: i + 1,
                        team_a_header: team1_player,
                        team_b_header: team2_player
                    }
                    players_data.append(row)
                
                df_players = pd.DataFrame(players_data)
                df_players.to_excel(writer, sheet_name=players_sheet, index=False)
                self._apply_sheet_styling(writer, players_sheet)
            
            # Sheet: Goals Summary - Count all goals from any action with outcome "ประตู"
            # Goals can come from: ยิง, ฟรีคิก, เตะมุม, ลูกโทษ, etc.
            all_goals = [e for e in self.events if e.outcome == "ประตู"]
            if all_goals:
                goals_data = []
                for goal in sorted(all_goals, key=lambda x: x.timestamp):
                    minutes = int(goal.timestamp // 60)
                    seconds = int(goal.timestamp % 60)
                    
                    if current_lang == "EN":
                        goals_data.append({
                            "Time": f"{minutes:02d}:{seconds:02d}",
                            "Minute": f"{int(goal.timestamp // 60)}",
                            "Type": self._t(goal.event_type, goal.event_type),
                            "Team": goal.team if goal.team else "Not Specified",
                            "Scorer": goal.player_name if goal.player_name else (f"#{goal.player_number}" if goal.player_number else "Not Specified"),
                            "Half": half_mapping.get(goal.half, f"Half {goal.half}"),
                            "Description": goal.description if goal.description else '-'
                        })
                    else:  # TH
                        goals_data.append({
                            "เวลา": f"{minutes:02d}:{seconds:02d}",
                            "นาที": f"{int(goal.timestamp // 60)}",
                            "ประเภท": self._t(goal.event_type, goal.event_type),
                            "ทีม": goal.team if goal.team else "ไม่ระบุ",
                            "ผู้ยิง": goal.player_name if goal.player_name else (f"#{goal.player_number}" if goal.player_number else "ไม่ระบุ"),
                            "ครึ่ง": half_mapping.get(goal.half, f"ครึ่ง {goal.half}"),
                            "คำอธิบาย": goal.description if goal.description else '-'
                        })
                
                df_goals = pd.DataFrame(goals_data)
                df_goals.to_excel(writer, sheet_name=goals_sheet, index=False)
                self._apply_sheet_styling(writer, goals_sheet)
            
            # Sheets: By Half
            for half_num, half_name in half_sheet_names.items():
                half_events = [e for e in self.events if e.half == half_num]
                if half_events:
                    summary_half = self._generate_summary(half_events)
                    summary_half.to_excel(writer, sheet_name=half_name, index=False)
                    self._apply_sheet_styling(writer, half_name)
            
            # Add professional statistics sheet with formulas (already has styling)
            self._add_professional_statistics_sheet(writer)
            
            # Add additional analysis sheets
            self._add_timeline_sheet(writer)
            self._apply_sheet_styling(writer, timeline_sheet)
            
            self._add_key_moments_sheet(writer)
            self._apply_sheet_styling(writer, key_moments_sheet)
            
            self._add_event_frequency_sheet(writer)
            self._apply_sheet_styling(writer, event_freq_sheet)
            
            self._add_set_pieces_sheet(writer)
            self._apply_sheet_styling(writer, set_pieces_sheet)
            
            if self.team_a_players or self.team_b_players:
                self._add_player_performance_sheet(writer)
                self._apply_sheet_styling(writer, player_stats_sheet)
    
    def _get_success_outcomes(self, event_type: str) -> List[str]:
        """
        Get list of successful outcomes for a specific event type
        
        คำอธิบาย:
        - "ประตู" = ได้ประตู (Goal)
        - "ยิงเข้า" = ยิงเข้าเป้าแต่ไม่ได้ประตู (Shot on Target - นับเป็นสำเร็จในการยิง)
        - "แอสซิสต์" = ส่งบอลแล้วได้ประตู (Assist - ส่งแล้วได้ประตู)
        - "คีย์พาส" = ส่งบอลที่สร้างโอกาสยิง (Key Pass - ส่งแล้วมีโอกาสยิง - นับเป็นสำเร็จ)
        - "สำเร็จ" = การกระทำสำเร็จ
        """
        success_mapping = {
            # Attacking Actions
            "ส่งบอล": ["สำเร็จ", "แอสซิสต์", "คีย์พาส"],
            # "แอสซิสต์" = ส่งแล้วได้ประตู, "คีย์พาส" = ส่งแล้วมีโอกาสยิง
            
            "ข้ามบอล": ["สำเร็จ", "แอสซิสต์", "คีย์พาส"],
            "ผ่านบอล": ["สำเร็จ", "แอสซิสต์", "คีย์พาส"],
            "ส่งบอลยาว": ["สำเร็จ", "แอสซิสต์", "คีย์พาส"],
            "ส่งบอลสั้น": ["สำเร็จ"],
            "ส่งบอลในเขตโทษ": ["สำเร็จ", "แอสซิสต์", "คีย์พาส"],
            
            "ยิง": ["ประตู", "ยิงเข้า"],
            # "ประตู" = ได้ประตู, "ยิงเข้า" = ยิงเข้าเป้าแต่ไม่ได้ประตู (นับเป็นสำเร็จ)
            # "ยิงออก", "บล็อก", "ถูกเซฟ" = ไม่สำเร็จ
            
            # Set Pieces
            "เตะมุม": ["ประตู", "ยิงเข้า", "แอสซิสต์", "คีย์พาส"],
            "ฟรีคิก": ["ประตู", "ยิงเข้า", "แอสซิสต์", "คีย์พาส"],
            "ลูกโทษ": ["ประตู"],
            # "ไม่ประตู", "ถูกเซฟ" = ไม่สำเร็จ
            
            "ทุ่มบอล": ["สำเร็จ"],
            
            # Defensive Actions
            "แย่งบอล": ["สำเร็จ"],
            "สกัดบอล": ["สำเร็จ"],
            "เคลียร์บอล": ["สำเร็จ"],
            "บล็อก": ["บล็อก", "บล็อกยิง"],
            "เซฟ": ["เซฟ", "เซฟสำคัญ"],
            
            # Disciplinary (ไม่มีสำเร็จ/ไม่สำเร็จ - เป็นเหตุการณ์)
            # หมายเหตุ: 
            # - "ฟาวล์" = ฟาวล์ธรรมดา (outcome: "ฟาวล์" หรือ None)
            #   ถ้า outcome เป็น "ใบเหลือง" หรือ "ใบแดง" จะถูกนับเป็นใบเหลือง/ใบแดง แทน
            # - "ใบเหลือง" = ใบเหลือง (อาจมาจาก action "ใบเหลือง" หรือ action "ฟาวล์" ที่ outcome เป็น "ใบเหลือง")
            # - "ใบแดง" = ใบแดง (อาจมาจาก action "ใบแดง" หรือ action "ฟาวล์" ที่ outcome เป็น "ใบแดง")
            "ฟาวล์": [],
            "ใบเหลือง": ["ใบเหลือง"],
            "ใบแดง": ["ใบแดง"],
            
            # Other Events
            "ออฟไซด์": [],
            "บอลออก": [],
            "เปลี่ยนตัว": ["เปลี่ยนตัวเข้า", "เปลี่ยนตัวออก"],
            "บาดเจ็บ": [],
            "เสียบอล": [],
            "ครองบอล": ["ครองบอล"],
        }
        return success_mapping.get(event_type, [])
    
    def _is_successful_outcome(self, event_type: str, outcome: str) -> bool:
        """Check if outcome is successful for specific event type"""
        if not outcome:
            return False
        success_outcomes = self._get_success_outcomes(event_type)
        return outcome in success_outcomes
    
    def _is_unsuccessful_outcome(self, event_type: str, outcome: str) -> bool:
        """
        Check if outcome is unsuccessful for specific event type
        
        Unsuccessful outcomes:
        - "ไม่สำเร็จ" = การกระทำไม่สำเร็จ
        - "ยิงออก" = ยิงออกนอกเป้า (Shot off Target)
        - "ถูกเซฟ" = ถูกผู้รักษาประตูเซฟ
        - "ไม่ประตู" = ไม่ได้ประตู (เช่น ลูกโทษ)
        - "ไม่เซฟ" = ไม่เซฟได้
        - "เสียบอล" = เสียบอล
        - "บล็อก" = ถูกบล็อก (เฉพาะในกรณี action "ยิง" เท่านั้น)
        """
        if not outcome:
            return False
        
        # สำหรับ action "บล็อก" ไม่มี unsuccessful (บล็อกสำเร็จ = "บล็อก" หรือ "บล็อกยิง")
        if event_type == "บล็อก":
            return False
        
        # Unsuccessful keywords
        unsuccessful_keywords = ["ไม่สำเร็จ", "ยิงออก", "ถูกเซฟ", "ไม่ประตู", "ไม่เซฟ", "เสียบอล"]
        
        # "บล็อก" = ถูกบล็อก (เฉพาะในกรณี action "ยิง", "เตะมุม", "ฟรีคิก")
        if outcome == "บล็อก" and event_type in ["ยิง", "เตะมุม", "ฟรีคิก"]:
            return True
        
        return any(keyword in outcome for keyword in unsuccessful_keywords)
    
    def _generate_summary(self, events: List[TrackingEvent]) -> pd.DataFrame:
        """Generate comprehensive summary statistics for events (without percentages)"""
        if not events:
            return pd.DataFrame()
        
        summary_data = []
        
        # Get all teams
        teams = sorted(set(e.team for e in events if e.team != "กลาง"))
        total_events = len(events)
        
        # === ส่วนที่ 1: สรุปภาพรวม ===
        summary_data.append({
            self._t('หมวดหมู่', 'หมวดหมู่'): self._t('ภาพรวม', 'ภาพรวม'),
            self._t('รายการ', 'รายการ'): self._t('จำนวนเหตุการณ์ทั้งหมด', 'จำนวนเหตุการณ์ทั้งหมด'),
            self._t('จำนวน', 'จำนวน'): total_events,
            self._t('หมายเหตุ', 'หมายเหตุ'): self._t('รวมทุกทีมและทุกประเภทเหตุการณ์', 'รวมทุกทีมและทุกประเภทเหตุการณ์')
        })
        
        summary_data.append({
            self._t('หมวดหมู่', 'หมวดหมู่'): self._t('ภาพรวม', 'ภาพรวม'),
            self._t('รายการ', 'รายการ'): self._t('จำนวนทีมที่วิเคราะห์', 'จำนวนทีมที่วิเคราะห์'),
            self._t('จำนวน', 'จำนวน'): len(teams),
            self._t('หมายเหตุ', 'หมายเหตุ'): ', '.join(teams) if teams else self._t('ไม่มีทีม', 'ไม่มีทีม')
        })
        
        # === ส่วนที่ 2: สรุปตามประเภทเหตุการณ์ ===
        event_stats = {}
        for event in events:
            event_type = event.event_type
            outcome = event.outcome if event.outcome else self._t("ไม่มีผลลัพธ์", "No Result")
            
            if event_type not in event_stats:
                event_stats[event_type] = {
                    'total': 0, 
                    'outcomes': {}, 
                    'by_team': {},
                    'successful': 0,
                    'unsuccessful': 0,
                    'attempts': 0  # สำหรับคำนวณอัตราความสำเร็จ (successful + unsuccessful)
                }
            
            event_stats[event_type]['total'] += 1
            
            # Count by outcome
            if outcome not in event_stats[event_type]['outcomes']:
                event_stats[event_type]['outcomes'][outcome] = {'total': 0, 'by_team': {}}
            event_stats[event_type]['outcomes'][outcome]['total'] += 1
            
            # Count successful/unsuccessful ตาม event type ที่ชัดเจน
            # หมายเหตุ: 
            # - "ประตู" = ได้ประตู (สำเร็จ)
            # - "ยิงเข้า" = ยิงเข้าเป้าแต่ไม่ได้ประตู (สำเร็จในการยิง)
            # - "แอสซิสต์" = ส่งแล้วได้ประตู (สำเร็จมาก)
            # - "คีย์พาส" = ส่งแล้วมีโอกาสยิง (สำเร็จ)
            # - "ยิงออก", "บล็อก", "ถูกเซฟ" = ไม่สำเร็จ (เฉพาะใน action "ยิง")
            if self._is_successful_outcome(event_type, outcome):
                event_stats[event_type]['successful'] += 1
                event_stats[event_type]['attempts'] += 1
            elif self._is_unsuccessful_outcome(event_type, outcome):
                event_stats[event_type]['unsuccessful'] += 1
                event_stats[event_type]['attempts'] += 1
            # หมายเหตุ: outcome ที่ไม่ใช่ success หรือ unsuccessful (เช่น "ออฟไซด์", "บอลออก") 
            # จะไม่นับใน attempts เพื่อความแม่นยำในการคำนวณอัตราความสำเร็จ
            
            # Count by team for this outcome
            team = event.team
            if team not in event_stats[event_type]['outcomes'][outcome]['by_team']:
                event_stats[event_type]['outcomes'][outcome]['by_team'][team] = 0
            event_stats[event_type]['outcomes'][outcome]['by_team'][team] += 1
            
            # Count by team for event type
            if team not in event_stats[event_type]['by_team']:
                event_stats[event_type]['by_team'][team] = 0
            event_stats[event_type]['by_team'][team] += 1
        
        # Add event type summaries
        for event_type, stats in sorted(event_stats.items()):
            event_total = stats['total']
            # Translate team names
            team_counts_list = []
            for team, count in sorted(stats['by_team'].items()):
                if team == "กลาง":
                    translated_team = self._t("กลาง", "Neutral")
                else:
                    translated_team = team  # Keep actual team name
                team_counts_list.append(f"{translated_team}: {count}")
            team_counts = ", ".join(team_counts_list)
            translated_event_type = self._t(event_type, event_type)
            
            summary_data.append({
                self._t('หมวดหมู่', 'หมวดหมู่'): self._t('ประเภทเหตุการณ์', 'ประเภทเหตุการณ์'),
                self._t('รายการ', 'รายการ'): translated_event_type,
                self._t('จำนวน', 'จำนวน'): event_total,
                self._t('หมายเหตุ', 'หมายเหตุ'): f"{self._t('แบ่งตามทีม:', 'แบ่งตามทีม:')} {team_counts}" if team_counts else self._t('ไม่มีข้อมูลทีม', 'ไม่มีข้อมูลทีม')
            })
            
            # Add outcome details
            for outcome, outcome_data in sorted(stats['outcomes'].items()):
                outcome_count = outcome_data['total']
                # Translate team names
                team_breakdown_list = []
                for team, count in sorted(outcome_data['by_team'].items()):
                    if team == "กลาง":
                        translated_team = self._t("กลาง", "Neutral")
                    else:
                        translated_team = team  # Keep actual team name
                    team_breakdown_list.append(f"{translated_team}: {count}")
                team_breakdown = ", ".join(team_breakdown_list)
                translated_outcome = self._t(outcome, outcome)
                
                summary_data.append({
                    self._t('หมวดหมู่', 'หมวดหมู่'): f'  └─ {translated_event_type}',
                    self._t('รายการ', 'รายการ'): f"{self._t('ผลลัพธ์:', 'ผลลัพธ์:')} {translated_outcome}",
                    self._t('จำนวน', 'จำนวน'): outcome_count,
                    self._t('หมายเหตุ', 'หมายเหตุ'): f"{self._t('แบ่งตามทีม:', 'แบ่งตามทีม:')} {team_breakdown}" if team_breakdown else '-'
                })
            
            # Add success rate info (คำนวณจาก attempts ที่มีผลสำเร็จ/ไม่สำเร็จเท่านั้น)
            attempts = stats['attempts']  # successful + unsuccessful
            if attempts > 0:
                success_rate = (stats['successful'] / attempts * 100) if attempts > 0 else 0
                success_text = self._t("สำเร็จ", "Success")
                failed_text = self._t("ไม่สำเร็จ", "Failed")
                rate_text = self._t("อัตรา", "Rate")
                summary_data.append({
                    self._t('หมวดหมู่', 'หมวดหมู่'): f'  └─ {translated_event_type}',
                    self._t('รายการ', 'รายการ'): self._t('อัตราความสำเร็จ', 'Success Rate'),
                    self._t('จำนวน', 'จำนวน'): f"{stats['successful']}/{attempts}",
                    self._t('หมายเหตุ', 'หมายเหตุ'): f'{success_text}: {stats["successful"]}, {failed_text}: {stats["unsuccessful"]}, {rate_text}: {success_rate:.1f}%'
                })
            elif stats['successful'] > 0:
                # กรณีที่ไม่มี unsuccessful แต่มี successful (เช่น ประตู)
                success_text = self._t("สำเร็จ", "Success")
                summary_data.append({
                    self._t('หมวดหมู่', 'หมวดหมู่'): f'  └─ {translated_event_type}',
                    self._t('รายการ', 'รายการ'): self._t('อัตราความสำเร็จ', 'Success Rate'),
                    self._t('จำนวน', 'จำนวน'): f"{stats['successful']}/{stats['successful']}",
                    self._t('หมายเหตุ', 'หมายเหตุ'): f'{success_text}: {stats["successful"]} (100%)'
                })
        
        # === ส่วนที่ 3: สรุปตามทีม ===
        for team in teams:
            team_events = [e for e in events if e.team == team]
            team_total = len(team_events)
            
            summary_data.append({
                self._t('หมวดหมู่', 'หมวดหมู่'): self._t('สรุปตามทีม', 'Team Summary'),
                self._t('รายการ', 'รายการ'): f'{team} - {self._t("รวมทั้งหมด", "Total")}',
                self._t('จำนวน', 'จำนวน'): team_total,
                self._t('หมายเหตุ', 'หมายเหตุ'): f'{self._t("จากทั้งหมด", "From Total")} {total_events} {self._t("เหตุการณ์", "Events")}'
            })
            
            # Team events by type
            team_event_types = {}
            team_outcomes = {}
            for event in team_events:
                # Count by type
                if event.event_type not in team_event_types:
                    team_event_types[event.event_type] = 0
                team_event_types[event.event_type] += 1
                
                # Count outcomes
                outcome = event.outcome if event.outcome else self._t("ไม่มีผลลัพธ์", "No Result")
                if outcome not in team_outcomes:
                    team_outcomes[outcome] = 0
                team_outcomes[outcome] += 1
            
            # Add team event types
            for event_type, count in sorted(team_event_types.items()):
                # Calculate success rate for this team and event type
                team_events_of_type = [e for e in team_events if e.event_type == event_type]
                success_outcomes = self._get_success_outcomes(event_type)
                # Translate unsuccessful keywords for matching
                unsuccessful_keywords_th = ["ไม่สำเร็จ", "ยิงออก", "ถูกเซฟ", "ไม่ประตู", "ไม่เซฟ", "เสียบอล"]
                unsuccessful_keywords_en = ["Failed", "Shot off Target", "Saved", "No Goal", "No Save", "Lost Ball"]
                # Use both Thai and English keywords for matching
                unsuccessful_keywords = unsuccessful_keywords_th + unsuccessful_keywords_en
                
                team_successful = len([e for e in team_events_of_type 
                                      if e.outcome and e.outcome in success_outcomes])
                team_unsuccessful = len([e for e in team_events_of_type 
                                        if e.outcome and self._is_unsuccessful_outcome(event_type, e.outcome)])
                team_attempts = team_successful + team_unsuccessful
                
                translated_event_type = self._t(event_type, event_type)
                if team_attempts > 0:
                    team_rate = (team_successful / team_attempts * 100)
                    success_text = self._t("สำเร็จ", "Success")
                    from_team_text = self._t("จาก", "From")
                    team_events_text = self._t("เหตุการณ์ของทีม", "Team Events")
                    summary_data.append({
                        self._t('หมวดหมู่', 'หมวดหมู่'): f'  └─ {team}',
                        self._t('รายการ', 'รายการ'): f'{translated_event_type} ({self._t("อัตราความสำเร็จ", "Success Rate")}: {team_rate:.1f}%)',
                        self._t('จำนวน', 'จำนวน'): count,
                        self._t('หมายเหตุ', 'หมายเหตุ'): f'{success_text}: {team_successful}/{team_attempts}, {from_team_text} {team_total} {team_events_text}'
                    })
                else:
                    from_team_text = self._t("จาก", "From")
                    team_events_text = self._t("เหตุการณ์ของทีม", "Team Events")
                    summary_data.append({
                        self._t('หมวดหมู่', 'หมวดหมู่'): f'  └─ {team}',
                        self._t('รายการ', 'รายการ'): translated_event_type,
                        self._t('จำนวน', 'จำนวน'): count,
                        self._t('หมายเหตุ', 'หมายเหตุ'): f'{from_team_text} {team_total} {team_events_text}'
                    })
        
        # === ส่วนที่ 4: สถิติเบื้องต้น ===
        # Calculate time distribution
        if events:
            times = [e.timestamp for e in events]
            min_time = min(times)
            max_time = max(times)
            total_duration = max_time - min_time
            total_minutes = total_duration / 60 if total_duration > 0 else 1
            
            seconds_text = self._t("วินาที", "seconds")
            summary_data.append({
                self._t('หมวดหมู่', 'หมวดหมู่'): self._t('สถิติเวลา', 'Time Statistics'),
                self._t('รายการ', 'รายการ'): self._t('เวลาตั้งแต่แรก', 'First Time'),
                self._t('จำนวน', 'จำนวน'): f"{int(min_time // 60)}:{int(min_time % 60):02d}",
                self._t('หมายเหตุ', 'หมายเหตุ'): f'{min_time:.1f} {seconds_text}'
            })
            
            summary_data.append({
                self._t('หมวดหมู่', 'หมวดหมู่'): self._t('สถิติเวลา', 'Time Statistics'),
                self._t('รายการ', 'รายการ'): self._t('เวลาสุดท้าย', 'Last Time'),
                self._t('จำนวน', 'จำนวน'): f"{int(max_time // 60)}:{int(max_time % 60):02d}",
                self._t('หมายเหตุ', 'หมายเหตุ'): f'{max_time:.1f} {seconds_text}'
            })
            
            minutes_text = self._t("นาที", "minutes")
            summary_data.append({
                self._t('หมวดหมู่', 'หมวดหมู่'): self._t('สถิติเวลา', 'Time Statistics'),
                self._t('รายการ', 'รายการ'): self._t('ระยะเวลารวม', 'Total Duration'),
                self._t('จำนวน', 'จำนวน'): f"{int(total_duration // 60)}:{int(total_duration % 60):02d}",
                self._t('หมายเหตุ', 'หมายเหตุ'): f'{total_duration:.1f} {seconds_text} ({total_minutes:.2f} {minutes_text})'
            })
            
            # Average events per minute
            events_per_minute = total_events / total_minutes if total_minutes > 0 else 0
            avg_text = self._t("เฉลี่ย", "Average")
            events_per_min_text = self._t("เหตุการณ์/นาที", "Events/Minute")
            summary_data.append({
                self._t('หมวดหมู่', 'หมวดหมู่'): self._t('สถิติเวลา', 'Time Statistics'),
                self._t('รายการ', 'รายการ'): self._t('จำนวนเหตุการณ์ต่อนาที', 'Events per Minute'),
                self._t('จำนวน', 'จำนวน'): f"{events_per_minute:.2f}",
                self._t('หมายเหตุ', 'หมายเหตุ'): f'{avg_text} {events_per_minute:.2f} {events_per_min_text}'
            })
            
            # Calculate time intervals between events
            if len(times) > 1:
                intervals = [times[i+1] - times[i] for i in range(len(times)-1)]
                avg_interval = sum(intervals) / len(intervals) if intervals else 0
                avg_per_event_text = self._t("วินาทีต่อเหตุการณ์", "seconds per event")
                summary_data.append({
                    self._t('หมวดหมู่', 'หมวดหมู่'): self._t('สถิติเวลา', 'Time Statistics'),
                    self._t('รายการ', 'รายการ'): self._t('ช่วงเวลาระหว่างเหตุการณ์ (เฉลี่ย)', 'Average Interval Between Events'),
                    self._t('จำนวน', 'จำนวน'): f"{avg_interval:.1f} {seconds_text}",
                    self._t('หมายเหตุ', 'หมายเหตุ'): f'{avg_text} {avg_interval:.1f} {avg_per_event_text}'
                })
        
        # Calculate half distribution
        half_counts = {}
        for event in events:
            if event.half not in half_counts:
                half_counts[event.half] = 0
            half_counts[event.half] += 1
        
        half_names = {
            1: self._t("ครึ่งแรก", "First Half"),
            2: self._t("ครึ่งหลัง", "Second Half"),
            3: self._t("ต่อเวลาครึ่งแรก", "Extra Time First Half"),
            4: self._t("ต่อเวลาครึ่งหลัง", "Extra Time Second Half")
        }
        
        from_total_text = self._t("จากทั้งหมด", "From Total")
        events_text = self._t("เหตุการณ์", "Events")
        for half_num, count in sorted(half_counts.items()):
            half_name = half_names.get(half_num, f"{self._t('ครึ่ง', 'Half')} {half_num}")
            summary_data.append({
                self._t('หมวดหมู่', 'หมวดหมู่'): self._t('สถิติตามครึ่ง', 'Half Statistics'),
                self._t('รายการ', 'รายการ'): half_name,
                self._t('จำนวน', 'จำนวน'): count,
                self._t('หมายเหตุ', 'หมายเหตุ'): f'{from_total_text} {total_events} {events_text}'
            })
        
        df_summary = pd.DataFrame(summary_data)
        # Translate column headers if DataFrame is not empty
        if not df_summary.empty and len(df_summary.columns) > 0:
            # Get translated column names
            translated_columns = []
            for col in df_summary.columns:
                translated_columns.append(self._t(col, col))
            df_summary.columns = translated_columns
        return df_summary
    
    def _generate_comparison(self, events: List[TrackingEvent]) -> pd.DataFrame:
        """Generate team comparison sheet (only when two teams exist)"""
        if not events:
            return pd.DataFrame()
        
        # Get all teams (excluding neutral)
        teams = sorted(set(e.team for e in events if e.team != "กลาง"))
        
        # Only generate comparison if exactly 2 teams
        if len(teams) != 2:
            return pd.DataFrame()
        
        team_a, team_b = teams[0], teams[1]
        team_a_events = [e for e in events if e.team == team_a]
        team_b_events = [e for e in events if e.team == team_b]
        
        comparison_data = []
        
        # === ส่วนที่ 1: เปรียบเทียบภาพรวม ===
        equal_text = self._t("เท่ากัน", "Equal")
        comparison_data.append({
            self._t('ตัวแปร', 'Variable'): self._t('จำนวนเหตุการณ์ทั้งหมด', 'Total Events'),
            team_a: len(team_a_events),
            team_b: len(team_b_events),
            self._t('ความแตกต่าง', 'Difference'): len(team_a_events) - len(team_b_events),
            self._t('ทีมที่มากกว่า', 'Team with More'): team_a if len(team_a_events) > len(team_b_events) else (team_b if len(team_b_events) > len(team_a_events) else equal_text)
        })
        
        # === ส่วนที่ 2: เปรียบเทียบตามประเภทเหตุการณ์ ===
        # Get all event types
        all_event_types = sorted(set(e.event_type for e in events))
        
        for event_type in all_event_types:
            team_a_count = len([e for e in team_a_events if e.event_type == event_type])
            team_b_count = len([e for e in team_b_events if e.event_type == event_type])
            diff = team_a_count - team_b_count
            translated_event_type = self._t(event_type, event_type)
            
            comparison_data.append({
                self._t('ตัวแปร', 'Variable'): f'{self._t("จำนวน", "Number of")} {translated_event_type}',
                team_a: team_a_count,
                team_b: team_b_count,
                self._t('ความแตกต่าง', 'Difference'): diff,
                self._t('ทีมที่มากกว่า', 'Team with More'): team_a if diff > 0 else (team_b if diff < 0 else equal_text)
            })
            
            # Compare outcomes for this event type
            team_a_outcomes = {}
            team_b_outcomes = {}
            
            for e in team_a_events:
                if e.event_type == event_type:
                    outcome = e.outcome if e.outcome else "ไม่มีผลลัพธ์"
                    team_a_outcomes[outcome] = team_a_outcomes.get(outcome, 0) + 1
            
            for e in team_b_events:
                if e.event_type == event_type:
                    outcome = e.outcome if e.outcome else "ไม่มีผลลัพธ์"
                    team_b_outcomes[outcome] = team_b_outcomes.get(outcome, 0) + 1
            
            # Compare each outcome
            all_outcomes = sorted(set(list(team_a_outcomes.keys()) + list(team_b_outcomes.keys())))
            for outcome in all_outcomes:
                a_count = team_a_outcomes.get(outcome, 0)
                b_count = team_b_outcomes.get(outcome, 0)
                diff = a_count - b_count
                
                comparison_data.append({
                    'ตัวแปร': f'  └─ {event_type} - {outcome}',
                    team_a: a_count,
                    team_b: b_count,
                    'ความแตกต่าง': diff,
                    'ทีมที่มากกว่า': team_a if diff > 0 else (team_b if diff < 0 else 'เท่ากัน')
                })
        
        # === ส่วนที่ 3: เปรียบเทียบอัตราความสำเร็จ ===
        # Calculate success rates for each team (ใช้ mapping ที่ชัดเจนและสอดคล้องกับ _is_successful_outcome/_is_unsuccessful_outcome)
        for event_type in all_event_types:
            # Team A
            team_a_events_of_type = [e for e in team_a_events if e.event_type == event_type]
            team_a_successful = len([e for e in team_a_events_of_type 
                                    if e.outcome and self._is_successful_outcome(event_type, e.outcome)])
            team_a_unsuccessful = len([e for e in team_a_events_of_type 
                                      if e.outcome and self._is_unsuccessful_outcome(event_type, e.outcome)])
            team_a_attempts = team_a_successful + team_a_unsuccessful
            team_a_rate = (team_a_successful / team_a_attempts * 100) if team_a_attempts > 0 else 0
            
            # Team B
            team_b_events_of_type = [e for e in team_b_events if e.event_type == event_type]
            team_b_successful = len([e for e in team_b_events_of_type 
                                    if e.outcome and self._is_successful_outcome(event_type, e.outcome)])
            team_b_unsuccessful = len([e for e in team_b_events_of_type 
                                      if e.outcome and self._is_unsuccessful_outcome(event_type, e.outcome)])
            team_b_attempts = team_b_successful + team_b_unsuccessful
            team_b_rate = (team_b_successful / team_b_attempts * 100) if team_b_attempts > 0 else 0
            
            # Only show if there are attempts (successful or unsuccessful)
            if team_a_attempts > 0 or team_b_attempts > 0:
                translated_event_type = self._t(event_type, event_type)
                comparison_data.append({
                    self._t('ตัวแปร', 'Variable'): f'{self._t("อัตราความสำเร็จ", "Success Rate")} {translated_event_type}',
                    team_a: f"{team_a_successful}/{team_a_attempts} ({team_a_rate:.1f}%)" if team_a_attempts > 0 else "-",
                    team_b: f"{team_b_successful}/{team_b_attempts} ({team_b_rate:.1f}%)" if team_b_attempts > 0 else "-",
                    self._t('ความแตกต่าง', 'Difference'): f"{team_a_rate - team_b_rate:.1f}%" if team_a_attempts > 0 and team_b_attempts > 0 else "-",
                    self._t('ทีมที่มากกว่า', 'Team with More'): team_a if team_a_rate > team_b_rate else (team_b if team_b_rate > team_a_rate else equal_text) if team_a_attempts > 0 and team_b_attempts > 0 else '-'
                })
        
        # === ส่วนที่ 4: เปรียบเทียบตามครึ่ง ===
        half_names = {
            1: self._t("ครึ่งแรก", "First Half"),
            2: self._t("ครึ่งหลัง", "Second Half"),
            3: self._t("ต่อเวลาครึ่งแรก", "Extra Time First Half"),
            4: self._t("ต่อเวลาครึ่งหลัง", "Extra Time Second Half")
        }
        
        all_halves = sorted(set(e.half for e in events))
        for half_num in all_halves:
            team_a_half = len([e for e in team_a_events if e.half == half_num])
            team_b_half = len([e for e in team_b_events if e.half == half_num])
            diff = team_a_half - team_b_half
            half_name = half_names.get(half_num, f"{self._t('ครึ่ง', 'Half')} {half_num}")
            
            comparison_data.append({
                self._t('ตัวแปร', 'Variable'): f'{self._t("จำนวนเหตุการณ์ใน", "Events in")} {half_name}',
                team_a: team_a_half,
                team_b: team_b_half,
                self._t('ความแตกต่าง', 'Difference'): diff,
                self._t('ทีมที่มากกว่า', 'Team with More'): team_a if diff > 0 else (team_b if diff < 0 else equal_text)
            })
        
        # === ส่วนที่ 5: สถิติเปรียบเทียบ ===
        # Average events per minute
        if events:
            times = [e.timestamp for e in events]
            total_time = max(times) - min(times) if times else 1
            minutes = total_time / 60 if total_time > 0 else 1
            
            team_a_per_min = len(team_a_events) / minutes if minutes > 0 else 0
            team_b_per_min = len(team_b_events) / minutes if minutes > 0 else 0
            
            comparison_data.append({
                self._t('ตัวแปร', 'Variable'): self._t('จำนวนเหตุการณ์ต่อนาที', 'Events per Minute'),
                team_a: f"{team_a_per_min:.2f}",
                team_b: f"{team_b_per_min:.2f}",
                self._t('ความแตกต่าง', 'Difference'): f"{team_a_per_min - team_b_per_min:.2f}",
                self._t('ทีมที่มากกว่า', 'Team with More'): team_a if team_a_per_min > team_b_per_min else (team_b if team_b_per_min > team_a_per_min else equal_text)
            })
            
            # Calculate overall success rate (across all event types)
            team_a_total_successful = 0
            team_a_total_attempts = 0
            team_b_total_successful = 0
            team_b_total_attempts = 0
            
            for event_type in all_event_types:
                # ใช้ logic เดียวกันกับ _is_successful_outcome/_is_unsuccessful_outcome เพื่อความสอดคล้อง
                team_a_events_of_type = [e for e in team_a_events if e.event_type == event_type]
                team_a_successful = len([e for e in team_a_events_of_type 
                                        if e.outcome and self._is_successful_outcome(event_type, e.outcome)])
                team_a_unsuccessful = len([e for e in team_a_events_of_type 
                                          if e.outcome and self._is_unsuccessful_outcome(event_type, e.outcome)])
                
                team_b_events_of_type = [e for e in team_b_events if e.event_type == event_type]
                team_b_successful = len([e for e in team_b_events_of_type 
                                        if e.outcome and self._is_successful_outcome(event_type, e.outcome)])
                team_b_unsuccessful = len([e for e in team_b_events_of_type 
                                          if e.outcome and self._is_unsuccessful_outcome(event_type, e.outcome)])
                
                team_a_total_successful += team_a_successful
                team_a_total_attempts += (team_a_successful + team_a_unsuccessful)
                team_b_total_successful += team_b_successful
                team_b_total_attempts += (team_b_successful + team_b_unsuccessful)
            
            if team_a_total_attempts > 0 or team_b_total_attempts > 0:
                team_a_overall_rate = (team_a_total_successful / team_a_total_attempts * 100) if team_a_total_attempts > 0 else 0
                team_b_overall_rate = (team_b_total_successful / team_b_total_attempts * 100) if team_b_total_attempts > 0 else 0
                
                comparison_data.append({
                    self._t('ตัวแปร', 'Variable'): self._t('อัตราความสำเร็จรวม (ทุกประเภท)', 'Overall Success Rate (All Types)'),
                    team_a: f"{team_a_total_successful}/{team_a_total_attempts} ({team_a_overall_rate:.1f}%)" if team_a_total_attempts > 0 else "-",
                    team_b: f"{team_b_total_successful}/{team_b_total_attempts} ({team_b_overall_rate:.1f}%)" if team_b_total_attempts > 0 else "-",
                    self._t('ความแตกต่าง', 'Difference'): f"{team_a_overall_rate - team_b_overall_rate:.1f}%" if team_a_total_attempts > 0 and team_b_total_attempts > 0 else "-",
                    self._t('ทีมที่มากกว่า', 'Team with More'): team_a if team_a_overall_rate > team_b_overall_rate else (team_b if team_b_overall_rate > team_a_overall_rate else equal_text) if team_a_total_attempts > 0 and team_b_total_attempts > 0 else '-'
                })
        
        # Most common event type for each team
        team_a_event_counts = {}
        team_b_event_counts = {}
        
        for e in team_a_events:
            team_a_event_counts[e.event_type] = team_a_event_counts.get(e.event_type, 0) + 1
        
        for e in team_b_events:
            team_b_event_counts[e.event_type] = team_b_event_counts.get(e.event_type, 0) + 1
        
        team_a_most_common = max(team_a_event_counts.items(), key=lambda x: x[1])[0] if team_a_event_counts else '-'
        team_b_most_common = max(team_b_event_counts.items(), key=lambda x: x[1])[0] if team_b_event_counts else '-'
        
        # Translate most common event types
        team_a_most_common_translated = self._t(team_a_most_common, team_a_most_common) if team_a_most_common != '-' else '-'
        team_b_most_common_translated = self._t(team_b_most_common, team_b_most_common) if team_b_most_common != '-' else '-'
        
        comparison_data.append({
            self._t('ตัวแปร', 'Variable'): self._t('ประเภทเหตุการณ์ที่เกิดขึ้นบ่อยที่สุด', 'Most Common Event Type'),
            team_a: team_a_most_common_translated,
            team_b: team_b_most_common_translated,
            self._t('ความแตกต่าง', 'Difference'): '-',
            self._t('ทีมที่มากกว่า', 'Team with More'): '-'
        })
        
        df_comparison = pd.DataFrame(comparison_data)
        # Translate column headers if DataFrame is not empty
        if not df_comparison.empty and len(df_comparison.columns) > 0:
            # Get translated column names
            translated_columns = []
            for col in df_comparison.columns:
                translated_columns.append(self._t(col, col))
            df_comparison.columns = translated_columns
        return df_comparison
    
    def _apply_sheet_styling(self, writer, sheet_name: str):
        """Apply beautiful styling to a sheet: auto-adjust columns, colors, borders, etc."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = writer.book
            if sheet_name not in wb.sheetnames:
                return
            
            ws = wb[sheet_name]
            
            # Style definitions
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            
            # Style header row (row 1)
            if ws.max_row > 0:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = border
                
                # Apply alternating row colors and borders to data rows
                light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                
                for row in range(2, ws.max_row + 1):
                    is_even = (row % 2 == 0)
                    fill_color = light_fill if is_even else white_fill
                    
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = fill_color
                        cell.border = border
                        
                        # Auto-align based on data type
                        if cell.value is not None:
                            # Check if cell contains formula
                            if cell.data_type == 'f':  # Formula
                                cell.alignment = center_align
                            elif isinstance(cell.value, (int, float)):
                                cell.alignment = center_align
                            else:
                                cell.alignment = left_align
            
            # Auto-adjust column widths
            for col in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                # Check header
                header_cell = ws.cell(row=1, column=col)
                if header_cell.value:
                    max_length = len(str(header_cell.value))
                
                # Check data cells (limit to first 100 rows for performance)
                for row in range(2, min(ws.max_row + 1, 102)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)
                        except:
                            pass
                
                # Set width with some padding (min 10, max 50)
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
        except Exception as e:
            # If styling fails, just continue without it
            pass
    
    def _add_timeline_sheet(self, writer):
        """Add timeline sheet showing all events in chronological order"""
        if not self.events:
            return
        
        timeline_data = []
        for event in sorted(self.events, key=lambda x: x.timestamp):
            minutes = int(event.timestamp // 60)
            seconds = int(event.timestamp % 60)
            half_names = {
                1: "ครึ่งแรก",
                2: "ครึ่งหลัง",
                3: "ต่อเวลาครึ่งแรก",
                4: "ต่อเวลาครึ่งหลัง"
            }
            
            timeline_data.append({
                self._t('เวลา', 'เวลา'): f"{minutes:02d}:{seconds:02d}",
                self._t('เวลา (วินาที)', 'เวลา (วินาที)'): event.timestamp,
                self._t('ครึ่ง', 'ครึ่ง'): half_names.get(event.half, f"{self._t('ครึ่ง', 'ครึ่ง')} {event.half}"),
                self._t('เหตุการณ์', 'เหตุการณ์'): self._t(event.event_type, event.event_type),
                self._t('ผลลัพธ์', 'ผลลัพธ์'): self._t(event.outcome, event.outcome) if event.outcome else '-',
                self._t('ทีม', 'ทีม'): self._t(event.team, event.team) if event.team else '-',
                self._t('หมายเลขผู้เล่น', 'หมายเลขผู้เล่น'): event.player_number if event.player_number else '-',
                self._t('ชื่อผู้เล่น', 'ชื่อผู้เล่น'): event.player_name if event.player_name else '-',
                self._t('คำอธิบาย', 'คำอธิบาย'): event.description if event.description else '-'
            })
        
        df_timeline = pd.DataFrame(timeline_data)
        timeline_sheet = self._t('ไทม์ไลน์', 'ไทม์ไลน์')
        df_timeline.to_excel(writer, sheet_name=timeline_sheet, index=False)
    
    def _add_key_moments_sheet(self, writer):
        """Add key moments sheet (goals, cards, substitutions)"""
        if not self.events:
            return
        
        key_moments = []
        
        # Goals - Count all goals from any action with outcome "ประตู"
        goals = [e for e in self.events if e.outcome == "ประตู"]
        for goal in sorted(goals, key=lambda x: x.timestamp):
            minutes = int(goal.timestamp // 60)
            seconds = int(goal.timestamp % 60)
            goal_type = f"{self._t('ประตู', 'ประตู')} ({self._t(goal.event_type, goal.event_type)})"  # Show action type
            key_moments.append({
                self._t('เวลา', 'เวลา'): f"{minutes:02d}:{seconds:02d}",
                self._t('ประเภท', 'ประเภท'): goal_type,
                self._t('ทีม', 'ทีม'): self._t(goal.team, goal.team) if goal.team else '-',
                self._t('ผู้เล่น', 'ผู้เล่น'): goal.player_name if goal.player_name else (f"#{goal.player_number}" if goal.player_number else '-'),
                self._t('คำอธิบาย', 'คำอธิบาย'): goal.description if goal.description else '-',
                self._t('ครึ่ง', 'ครึ่ง'): f"{self._t('ครึ่ง', 'ครึ่ง')} {goal.half}"
            })
        
        # Yellow cards
        yellow_cards = []
        all_yellow = [e for e in self.events if e.event_type == "ใบเหลือง"]
        all_yellow_from_fouls = [e for e in self.events if e.event_type == "ฟาวล์" and e.outcome == "ใบเหลือง"]
        yellow_cards = all_yellow + all_yellow_from_fouls
        
        for card in sorted(yellow_cards, key=lambda x: x.timestamp):
            minutes = int(card.timestamp // 60)
            seconds = int(card.timestamp % 60)
            key_moments.append({
                self._t('เวลา', 'เวลา'): f"{minutes:02d}:{seconds:02d}",
                self._t('ประเภท', 'ประเภท'): self._t('ใบเหลือง', 'ใบเหลือง'),
                self._t('ทีม', 'ทีม'): self._t(card.team, card.team) if card.team else '-',
                self._t('ผู้เล่น', 'ผู้เล่น'): card.player_name if card.player_name else (f"#{card.player_number}" if card.player_number else '-'),
                self._t('คำอธิบาย', 'คำอธิบาย'): card.description if card.description else '-',
                self._t('ครึ่ง', 'ครึ่ง'): f"{self._t('ครึ่ง', 'ครึ่ง')} {card.half}"
            })
        
        # Red cards
        red_cards = []
        all_red = [e for e in self.events if e.event_type == "ใบแดง"]
        all_red_from_fouls = [e for e in self.events if e.event_type == "ฟาวล์" and e.outcome == "ใบแดง"]
        red_cards = all_red + all_red_from_fouls
        
        for card in sorted(red_cards, key=lambda x: x.timestamp):
            minutes = int(card.timestamp // 60)
            seconds = int(card.timestamp % 60)
            key_moments.append({
                self._t('เวลา', 'เวลา'): f"{minutes:02d}:{seconds:02d}",
                self._t('ประเภท', 'ประเภท'): self._t('ใบแดง', 'ใบแดง'),
                self._t('ทีม', 'ทีม'): self._t(card.team, card.team) if card.team else '-',
                self._t('ผู้เล่น', 'ผู้เล่น'): card.player_name if card.player_name else (f"#{card.player_number}" if card.player_number else '-'),
                self._t('คำอธิบาย', 'คำอธิบาย'): card.description if card.description else '-',
                self._t('ครึ่ง', 'ครึ่ง'): f"{self._t('ครึ่ง', 'ครึ่ง')} {card.half}"
            })
        
        # Substitutions
        substitutions = [e for e in self.events if e.event_type == "เปลี่ยนตัว"]
        for sub in sorted(substitutions, key=lambda x: x.timestamp):
            minutes = int(sub.timestamp // 60)
            seconds = int(sub.timestamp % 60)
            key_moments.append({
                self._t('เวลา', 'เวลา'): f"{minutes:02d}:{seconds:02d}",
                self._t('ประเภท', 'ประเภท'): self._t('เปลี่ยนตัว', 'เปลี่ยนตัว'),
                self._t('ทีม', 'ทีม'): self._t(sub.team, sub.team) if sub.team else '-',
                self._t('ผู้เล่น', 'ผู้เล่น'): sub.player_name if sub.player_name else (f"#{sub.player_number}" if sub.player_number else '-'),
                self._t('คำอธิบาย', 'คำอธิบาย'): sub.description if sub.description else (self._t(sub.outcome, sub.outcome) if sub.outcome else '-'),
                self._t('ครึ่ง', 'ครึ่ง'): f"{self._t('ครึ่ง', 'ครึ่ง')} {sub.half}"
            })
        
        if key_moments:
            df_key_moments = pd.DataFrame(key_moments)
            key_moments_sheet = self._t('ช่วงเวลาสำคัญ', 'ช่วงเวลาสำคัญ')
            df_key_moments.to_excel(writer, sheet_name=key_moments_sheet, index=False)
    
    def _add_event_frequency_sheet(self, writer):
        """Add event frequency analysis by time periods"""
        if not self.events:
            return
        
        # Get time range
        times = [e.timestamp for e in self.events]
        if not times:
            return
        
        min_time = min(times)
        max_time = max(times)
        total_duration = max_time - min_time
        
        # Divide into 10-minute periods
        period_minutes = 10
        periods = []
        current_time = 0
        
        # Get all unique event types (limit to most common ones to avoid too many columns)
        all_event_types = sorted(set(e.event_type for e in self.events))
        # Focus on key event types
        key_event_types = ["ยิง", "ส่งบอล", "ฟาวล์", "ใบเหลือง", "ใบแดง", "แย่งบอล", "เตะมุม", "ฟรีคิก"]
        event_types_to_track = [et for et in key_event_types if et in all_event_types]
        # Add other event types if not too many
        if len(all_event_types) <= 15:
            event_types_to_track = all_event_types
        
        while current_time <= total_duration:
            period_start = current_time
            period_end = min(current_time + period_minutes * 60, total_duration)
            
            period_events = [e for e in self.events 
                           if period_start <= (e.timestamp - min_time) <= period_end]
            
            event_counts = {}
            for event in period_events:
                event_type = event.event_type
                event_counts[event_type] = event_counts.get(event_type, 0) + 1
            
            period_data = {
                'ช่วงเวลา': f"{int(period_start // 60)}:00 - {int(period_end // 60)}:00",
                'เวลาเริ่มต้น (นาที)': int(period_start // 60),
                'เวลาสิ้นสุด (นาที)': int(period_end // 60),
                'จำนวนเหตุการณ์ทั้งหมด': len(period_events)
            }
            
            # Add counts for tracked event types
            for et in event_types_to_track:
                translated_et = self._t(et, et)
                period_data[f'{self._t("จำนวน", "Number of")} {translated_et}'] = event_counts.get(et, 0)
            
            periods.append(period_data)
            
            current_time += period_minutes * 60
        
        if periods:
            df_frequency = pd.DataFrame(periods)
            df_frequency.to_excel(writer, sheet_name='ความถี่เหตุการณ์', index=False)
    
    def _add_set_pieces_sheet(self, writer):
        """Add set pieces summary (corners, free kicks, penalties)"""
        if not self.events:
            return
        
        set_piece_types = ["เตะมุม", "ฟรีคิก", "ลูกโทษ"]
        set_pieces = [e for e in self.events if e.event_type in set_piece_types]
        
        if not set_pieces:
            return
        
        set_pieces_data = []
        for event in sorted(set_pieces, key=lambda x: x.timestamp):
            minutes = int(event.timestamp // 60)
            seconds = int(event.timestamp % 60)
            half_names = {
                1: self._t("ครึ่งแรก", "First Half"),
                2: self._t("ครึ่งหลัง", "Second Half"),
                3: self._t("ต่อเวลาครึ่งแรก", "Extra Time First Half"),
                4: self._t("ต่อเวลาครึ่งหลัง", "Extra Time Second Half")
            }
            
            set_pieces_data.append({
                self._t('เวลา', 'เวลา'): f"{minutes:02d}:{seconds:02d}",
                self._t('ประเภท', 'ประเภท'): self._t(event.event_type, event.event_type),
                self._t('ผลลัพธ์', 'ผลลัพธ์'): self._t(event.outcome, event.outcome) if event.outcome else '-',
                self._t('ทีม', 'ทีม'): self._t(event.team, event.team) if event.team else '-',
                self._t('หมายเลขผู้เล่น', 'หมายเลขผู้เล่น'): event.player_number if event.player_number else '-',
                self._t('ชื่อผู้เล่น', 'ชื่อผู้เล่น'): event.player_name if event.player_name else '-',
                self._t('ครึ่ง', 'ครึ่ง'): half_names.get(event.half, f"{self._t('ครึ่ง', 'ครึ่ง')} {event.half}"),
                self._t('คำอธิบาย', 'คำอธิบาย'): event.description if event.description else '-'
            })
        
        df_set_pieces = pd.DataFrame(set_pieces_data)
        set_pieces_sheet = self._t('ลูกตั้งเตะ', 'ลูกตั้งเตะ')
        df_set_pieces.to_excel(writer, sheet_name=set_pieces_sheet, index=False)
    
    def _add_player_performance_sheet(self, writer):
        """Add player performance statistics"""
        if not self.events:
            return
        
        # Get teams
        teams = sorted(set(e.team for e in self.events if e.team and e.team != "กลาง"))
        if not teams:
            return
        
        player_stats = {}
        
        for event in self.events:
            if not event.player_number and not event.player_name:
                continue
            
            # Use player_number as key, fallback to player_name
            player_key = f"#{event.player_number}" if event.player_number else event.player_name
            team = event.team if event.team else self._t("ไม่ระบุ", "Not Specified")
            
            if player_key not in player_stats:
                player_stats[player_key] = {
                    self._t('ชื่อผู้เล่น', 'Player Name'): event.player_name if event.player_name else player_key,
                    self._t('หมายเลข', 'Number'): event.player_number if event.player_number else '-',
                    self._t('ทีม', 'Team'): team,
                    self._t('จำนวนเหตุการณ์', 'Total Events'): 0,
                    self._t('ประตู', 'Goals'): 0,
                    self._t('แอสซิสต์', 'Assists'): 0,
                    self._t('ใบเหลือง', 'Yellow Cards'): 0,
                    self._t('ใบแดง', 'Red Cards'): 0,
                    self._t('ฟาวล์', 'Fouls'): 0
                }
            
            player_stats[player_key][self._t('จำนวนเหตุการณ์', 'Total Events')] += 1
            
            # Count specific events
            # Count goals from any action with outcome "ประตู" (not just "ยิง")
            if event.outcome == "ประตู":
                player_stats[player_key][self._t('ประตู', 'Goals')] += 1
            elif event.outcome == "แอสซิสต์":
                player_stats[player_key][self._t('แอสซิสต์', 'Assists')] += 1
            elif event.event_type == "ใบเหลือง" or (event.event_type == "ฟาวล์" and event.outcome == "ใบเหลือง"):
                player_stats[player_key][self._t('ใบเหลือง', 'Yellow Cards')] += 1
            elif event.event_type == "ใบแดง" or (event.event_type == "ฟาวล์" and event.outcome == "ใบแดง"):
                player_stats[player_key][self._t('ใบแดง', 'Red Cards')] += 1
            elif event.event_type == "ฟาวล์":
                player_stats[player_key][self._t('ฟาวล์', 'Fouls')] += 1
        
        if player_stats:
            player_data = list(player_stats.values())
            df_players = pd.DataFrame(player_data)
            # Sort by team, then by goals, then by events
            team_col = self._t('ทีม', 'Team')
            goals_col = self._t('ประตู', 'Goals')
            events_col = self._t('จำนวนเหตุการณ์', 'Total Events')
            df_players = df_players.sort_values([team_col, goals_col, events_col], ascending=[True, False, False])
            player_stats_sheet = self._t('สถิติผู้เล่น', 'Player Statistics')
            df_players.to_excel(writer, sheet_name=player_stats_sheet, index=False)
    
    def _add_professional_statistics_sheet(self, writer):
        """Add professional statistics sheet with formulas for advanced analysis"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Get current language to ensure consistent translation
            current_lang = self._get_current_language()
            
            # Get workbook directly from writer (newer pandas API)
            wb = writer.book
            
            # Create new sheet - use language-specific names
            if current_lang == "EN":
                stats_sheet_name = "Analysis Statistics"
                match_analysis_text = "Match Analysis Statistics:"
                variable_text = "Variable"
                difference_text = "Difference"
                percentage_text = "Percentage (%)"
                notes_text = "Notes"
                cannot_create_text = "Cannot create comparison statistics (requires 2 teams)"
                equal_text = "Equal"
                shooting_stats_text = "1. Shooting Statistics"
                passing_stats_text = "2. Passing Statistics"
                defensive_stats_text = "3. Defensive Statistics"
                disciplinary_stats_text = "4. Disciplinary Statistics"
                attacking_stats_text = "5. Attacking Statistics"
                overall_summary_text = "6. Overall Summary"
            else:  # TH
                stats_sheet_name = "สถิติการวิเคราะห์"
                match_analysis_text = "สถิติการวิเคราะห์การแข่งขัน:"
                variable_text = "ตัวแปร"
                difference_text = "ความแตกต่าง"
                percentage_text = "เปอร์เซ็นต์ (%)"
                notes_text = "หมายเหตุ"
                cannot_create_text = "ไม่สามารถสร้างสถิติเปรียบเทียบได้ (ต้องมี 2 ทีม)"
                equal_text = "เท่ากัน"
                shooting_stats_text = "1. สถิติการยิง (Shooting)"
                passing_stats_text = "2. สถิติการส่งบอล (Passing)"
                defensive_stats_text = "3. สถิติการป้องกัน (Defensive)"
                disciplinary_stats_text = "4. สถิติการทำผิดกติกา (Disciplinary)"
                attacking_stats_text = "5. สถิติการโจมตี (Attacking)"
                overall_summary_text = "6. สรุปภาพรวม (Overall Summary)"
            
            ws = wb.create_sheet(stats_sheet_name, 0)  # Insert at beginning
            
            # Get teams
            teams = sorted(set(e.team for e in self.events if e.team != "กลาง"))
            if len(teams) != 2:
                # If not 2 teams, create basic stats
                ws['A1'] = stats_sheet_name
                ws['A2'] = cannot_create_text
                # Workbook will be saved automatically when context manager exits
                return
            
            team_a, team_b = teams[0], teams[1]
            # Filter events by team - use exact match and handle None/empty team values
            # Also handle case where team might be stored as string with different formatting
            # Use normalized string comparison (strip whitespace, case-insensitive)
            team_a_events = [e for e in self.events 
                           if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()]
            team_b_events = [e for e in self.events 
                           if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()]
            
            # Style definitions
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            title_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            
            # Title
            ws.merge_cells('A1:F1')
            ws['A1'] = f'{match_analysis_text} {self.match_name}'
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align
            
            # Team names row
            ws['A2'] = variable_text
            ws['B2'] = team_a
            ws['C2'] = team_b
            ws['D2'] = difference_text
            ws['E2'] = percentage_text
            ws['F2'] = notes_text
            
            for col in ['A2', 'B2', 'C2', 'D2', 'E2', 'F2']:
                ws[col].fill = header_fill
                ws[col].font = header_font
                ws[col].alignment = center_align
                ws[col].border = border
            
            row = 3
            
            # === 1. สถิติการยิง (Shooting Statistics) ===
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = shooting_stats_text
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            
            # Calculate shooting stats - get all shots first, then filter by team
            all_shots = [e for e in self.events if e.event_type == "ยิง"]
            team_a_shots = [e for e in all_shots 
                          if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()]
            team_b_shots = [e for e in all_shots 
                          if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()]
            
            team_a_goals = len([e for e in team_a_shots if e.outcome == "ประตู"])
            team_b_goals = len([e for e in team_b_shots if e.outcome == "ประตู"])
            
            team_a_shots_on_target = len([e for e in team_a_shots if e.outcome in ["ประตู", "ยิงเข้า"]])
            team_b_shots_on_target = len([e for e in team_b_shots if e.outcome in ["ประตู", "ยิงเข้า"]])
            
            team_a_shots_off_target = len([e for e in team_a_shots if e.outcome == "ยิงออก"])
            team_b_shots_off_target = len([e for e in team_b_shots if e.outcome == "ยิงออก"])
            
            team_a_blocked = len([e for e in team_a_shots if e.outcome == "บล็อก"])
            team_b_blocked = len([e for e in team_b_shots if e.outcome == "บล็อก"])
            
            team_a_saved = len([e for e in team_a_shots if e.outcome == "ถูกเซฟ"])
            team_b_saved = len([e for e in team_b_shots if e.outcome == "ถูกเซฟ"])
            
            # Define shooting stats based on language
            if current_lang == "EN":
                stats_shooting = [
                    ('Total Shots', len(team_a_shots), len(team_b_shots)),
                    ('Goals', team_a_goals, team_b_goals),
                    ('Shots on Target', team_a_shots_on_target, team_b_shots_on_target),
                    ('Shots off Target', team_a_shots_off_target, team_b_shots_off_target),
                    ('Blocked', team_a_blocked, team_b_blocked),
                    ('Saved', team_a_saved, team_b_saved),
                ]
                shot_on_target_rate_text = 'Shot on Target Rate (%)'
                goal_conversion_rate_text = 'Goal Conversion Rate (%)'
            else:  # TH
                stats_shooting = [
                    ('จำนวนการยิงทั้งหมด', len(team_a_shots), len(team_b_shots)),
                    ('จำนวนประตู', team_a_goals, team_b_goals),
                    ('ยิงเข้าเป้า', team_a_shots_on_target, team_b_shots_on_target),
                    ('ยิงออกนอกเป้า', team_a_shots_off_target, team_b_shots_off_target),
                    ('ถูกบล็อก', team_a_blocked, team_b_blocked),
                    ('ถูกเซฟ', team_a_saved, team_b_saved),
                ]
                shot_on_target_rate_text = 'อัตราการยิงเข้าเป้า (%)'
                goal_conversion_rate_text = 'อัตราการยิงได้ประตู (%)'
            
            # Store row numbers for later reference
            row_total_shots = row  # Row for "จำนวนการยิงทั้งหมด"
            row_goals = row + 1    # Row for "จำนวนประตู"
            
            for stat_name, val_a, val_b in stats_shooting:
                ws[f'A{row}'] = stat_name
                ws[f'B{row}'] = val_a
                ws[f'C{row}'] = val_b
                ws[f'D{row}'] = f'=B{row}-C{row}'
                if val_a + val_b > 0:
                    ws[f'E{row}'] = f'=IF(B{row}+C{row}>0, B{row}/(B{row}+C{row})*100, 0)'
                else:
                    ws[f'E{row}'] = 0
                ws[f'F{row}'] = f'=IF(D{row}>0, "{team_a}", IF(D{row}<0, "{team_b}", "{equal_text}"))'
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{row}'].border = border
                    if col in ['B', 'C', 'D', 'E']:
                        ws[f'{col}{row}'].alignment = center_align
                row += 1
            
            # Shooting accuracy formulas
            ws[f'A{row}'] = shot_on_target_rate_text
            ws[f'B{row}'] = f'=IF(B{row-6}>0, B{row-5}/B{row-6}*100, 0)'
            ws[f'C{row}'] = f'=IF(C{row-6}>0, C{row-5}/C{row-6}*100, 0)'
            ws[f'D{row}'] = f'=B{row}-C{row}'
            ws[f'E{row}'] = f'=IF(B{row}+C{row}>0, B{row}/(B{row}+C{row})*100, 0)'
            ws[f'F{row}'] = f'=IF(D{row}>0, "{team_a}", IF(D{row}<0, "{team_b}", "{equal_text}"))'
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row}'].border = border
                ws[f'{col}{row}'].font = Font(bold=True)
                if col in ['B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].alignment = center_align
            row += 1
            
            ws[f'A{row}'] = goal_conversion_rate_text
            ws[f'B{row}'] = f'=IF(B{row-7}>0, B{row-6}/B{row-7}*100, 0)'
            ws[f'C{row}'] = f'=IF(C{row-7}>0, C{row-6}/C{row-7}*100, 0)'
            ws[f'D{row}'] = f'=B{row}-C{row}'
            ws[f'E{row}'] = f'=IF(B{row}+C{row}>0, B{row}/(B{row}+C{row})*100, 0)'
            ws[f'F{row}'] = f'=IF(D{row}>0, "{team_a}", IF(D{row}<0, "{team_b}", "{equal_text}"))'
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row}'].border = border
                ws[f'{col}{row}'].font = Font(bold=True)
                if col in ['B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].alignment = center_align
            row += 2
            
            # === 2. สถิติการส่งบอล (Passing Statistics) ===
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = passing_stats_text
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            
            pass_types = ["ส่งบอล", "ข้ามบอล", "ผ่านบอล", "ส่งบอลยาว", "ส่งบอลสั้น", "ส่งบอลในเขตโทษ"]
            all_passes = [e for e in self.events if e.event_type in pass_types]
            team_a_passes = [e for e in all_passes 
                           if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()]
            team_b_passes = [e for e in all_passes 
                           if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()]
            
            team_a_passes_success = len([e for e in team_a_passes if e.outcome in ["สำเร็จ", "แอสซิสต์", "คีย์พาส"]])
            team_b_passes_success = len([e for e in team_b_passes if e.outcome in ["สำเร็จ", "แอสซิสต์", "คีย์พาส"]])
            
            team_a_passes_failed = len([e for e in team_a_passes if e.outcome == "ไม่สำเร็จ"])
            team_b_passes_failed = len([e for e in team_b_passes if e.outcome == "ไม่สำเร็จ"])
            
            team_a_assists = len([e for e in team_a_passes if e.outcome == "แอสซิสต์"])
            team_b_assists = len([e for e in team_b_passes if e.outcome == "แอสซิสต์"])
            
            team_a_key_passes = len([e for e in team_a_passes if e.outcome == "คีย์พาส"])
            team_b_key_passes = len([e for e in team_b_passes if e.outcome == "คีย์พาส"])
            
            # Define passing stats based on language
            if current_lang == "EN":
                stats_passing = [
                    ('Total Passes', len(team_a_passes), len(team_b_passes)),
                    ('Successful Passes', team_a_passes_success, team_b_passes_success),
                    ('Failed Passes', team_a_passes_failed, team_b_passes_failed),
                    ('Assists', team_a_assists, team_b_assists),
                    ('Key Passes', team_a_key_passes, team_b_key_passes),
                ]
                pass_success_rate_text = 'Pass Success Rate (%)'
            else:  # TH
                stats_passing = [
                    ('จำนวนการส่งบอลทั้งหมด', len(team_a_passes), len(team_b_passes)),
                    ('ส่งบอลสำเร็จ', team_a_passes_success, team_b_passes_success),
                    ('ส่งบอลไม่สำเร็จ', team_a_passes_failed, team_b_passes_failed),
                    ('แอสซิสต์', team_a_assists, team_b_assists),
                    ('คีย์พาส', team_a_key_passes, team_b_key_passes),
                ]
                pass_success_rate_text = 'อัตราการส่งบอลสำเร็จ (%)'
            
            for stat_name, val_a, val_b in stats_passing:
                ws[f'A{row}'] = stat_name
                ws[f'B{row}'] = val_a
                ws[f'C{row}'] = val_b
                ws[f'D{row}'] = f'=B{row}-C{row}'
                if val_a + val_b > 0:
                    ws[f'E{row}'] = f'=IF(B{row}+C{row}>0, B{row}/(B{row}+C{row})*100, 0)'
                else:
                    ws[f'E{row}'] = 0
                ws[f'F{row}'] = f'=IF(D{row}>0, "{team_a}", IF(D{row}<0, "{team_b}", "{equal_text}"))'
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{row}'].border = border
                    if col in ['B', 'C', 'D', 'E']:
                        ws[f'{col}{row}'].alignment = center_align
                row += 1
            
            # Pass accuracy formula
            ws[f'A{row}'] = pass_success_rate_text
            ws[f'B{row}'] = f'=IF(B{row-5}>0, B{row-4}/(B{row-4}+B{row-3})*100, 0)'
            ws[f'C{row}'] = f'=IF(C{row-5}>0, C{row-4}/(C{row-4}+C{row-3})*100, 0)'
            ws[f'D{row}'] = f'=B{row}-C{row}'
            ws[f'E{row}'] = f'=IF(B{row}+C{row}>0, B{row}/(B{row}+C{row})*100, 0)'
            ws[f'F{row}'] = f'=IF(D{row}>0, "{team_a}", IF(D{row}<0, "{team_b}", "{equal_text}"))'
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row}'].border = border
                ws[f'{col}{row}'].font = Font(bold=True)
                if col in ['B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].alignment = center_align
            row += 2
            
            # === 3. สถิติการป้องกัน (Defensive Statistics) ===
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = defensive_stats_text
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            
            defensive_types = ["แย่งบอล", "สกัดบอล", "เคลียร์บอล", "บล็อก", "เซฟ"]
            all_defensive = [e for e in self.events if e.event_type in defensive_types]
            team_a_defensive = [e for e in all_defensive 
                              if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()]
            team_b_defensive = [e for e in all_defensive 
                              if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()]
            
            all_tackles = [e for e in self.events if e.event_type == "แย่งบอล"]
            team_a_tackles = len([e for e in all_tackles 
                                if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()])
            team_b_tackles = len([e for e in all_tackles 
                                if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()])
            
            all_interceptions = [e for e in self.events if e.event_type == "สกัดบอล"]
            team_a_interceptions = len([e for e in all_interceptions 
                                      if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()])
            team_b_interceptions = len([e for e in all_interceptions 
                                      if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()])
            
            all_clearances = [e for e in self.events if e.event_type == "เคลียร์บอล"]
            team_a_clearances = len([e for e in all_clearances 
                                   if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()])
            team_b_clearances = len([e for e in all_clearances 
                                   if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()])
            
            all_blocks = [e for e in self.events if e.event_type == "บล็อก"]
            team_a_blocks = len([e for e in all_blocks 
                               if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()])
            team_b_blocks = len([e for e in all_blocks 
                               if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()])
            
            all_saves = [e for e in self.events if e.event_type == "เซฟ"]
            team_a_saves = len([e for e in all_saves 
                              if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()])
            team_b_saves = len([e for e in all_saves 
                              if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()])
            
            # Define defensive stats based on language
            if current_lang == "EN":
                stats_defensive = [
                    ('Tackles', team_a_tackles, team_b_tackles),
                    ('Interceptions', team_a_interceptions, team_b_interceptions),
                    ('Clearances', team_a_clearances, team_b_clearances),
                    ('Blocks', team_a_blocks, team_b_blocks),
                    ('Saves', team_a_saves, team_b_saves),
                    ('Total Defensive Actions', len(team_a_defensive), len(team_b_defensive)),
                ]
            else:  # TH
                stats_defensive = [
                    ('แย่งบอล', team_a_tackles, team_b_tackles),
                    ('สกัดบอล', team_a_interceptions, team_b_interceptions),
                    ('เคลียร์บอล', team_a_clearances, team_b_clearances),
                    ('บล็อก', team_a_blocks, team_b_blocks),
                    ('เซฟ', team_a_saves, team_b_saves),
                    ('รวมการป้องกัน', len(team_a_defensive), len(team_b_defensive)),
                ]
            
            for stat_name, val_a, val_b in stats_defensive:
                ws[f'A{row}'] = stat_name
                ws[f'B{row}'] = val_a
                ws[f'C{row}'] = val_b
                ws[f'D{row}'] = f'=B{row}-C{row}'
                if val_a + val_b > 0:
                    ws[f'E{row}'] = f'=IF(B{row}+C{row}>0, B{row}/(B{row}+C{row})*100, 0)'
                else:
                    ws[f'E{row}'] = 0
                ws[f'F{row}'] = f'=IF(D{row}>0, "{team_a}", IF(D{row}<0, "{team_b}", "{equal_text}"))'
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{row}'].border = border
                    if col in ['B', 'C', 'D', 'E']:
                        ws[f'{col}{row}'].alignment = center_align
                row += 1
            row += 1
            
            # === 4. สถิติการทำผิดกติกา (Disciplinary Statistics) ===
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = disciplinary_stats_text
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            
            # Count disciplinary events - handle overlapping actions and outcomes
            # Strategy:
            # - "ฟาวล์" = action "ฟาวล์" ที่ outcome เป็น "ฟาวล์" หรือ None (ไม่ใช่ใบเหลือง/ใบแดง)
            # - "ใบเหลือง" = action "ใบเหลือง" + action "ฟาวล์" ที่ outcome เป็น "ใบเหลือง"
            # - "ใบแดง" = action "ใบแดง" + action "ฟาวล์" ที่ outcome เป็น "ใบแดง"
            
            # Get all foul events
            all_foul_events = [e for e in self.events if e.event_type == "ฟาวล์"]
            # Get all yellow card events (both action "ใบเหลือง" and fouls with outcome "ใบเหลือง")
            all_yellow_events = [e for e in self.events if e.event_type == "ใบเหลือง"]
            all_yellow_from_fouls = [e for e in all_foul_events if e.outcome == "ใบเหลือง"]
            # Get all red card events (both action "ใบแดง" and fouls with outcome "ใบแดง")
            all_red_events = [e for e in self.events if e.event_type == "ใบแดง"]
            all_red_from_fouls = [e for e in all_foul_events if e.outcome == "ใบแดง"]
            
            # Count fouls (only fouls without yellow/red card outcomes)
            team_a_fouls = len([e for e in all_foul_events 
                              if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()
                              and e.outcome != "ใบเหลือง" and e.outcome != "ใบแดง"])
            team_b_fouls = len([e for e in all_foul_events 
                              if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()
                              and e.outcome != "ใบเหลือง" and e.outcome != "ใบแดง"])
            
            # Count yellow cards (action "ใบเหลือง" + fouls with outcome "ใบเหลือง")
            all_yellow_combined = all_yellow_events + all_yellow_from_fouls
            team_a_yellow = len([e for e in all_yellow_combined 
                               if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()])
            team_b_yellow = len([e for e in all_yellow_combined 
                               if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()])
            
            # Count red cards (action "ใบแดง" + fouls with outcome "ใบแดง")
            all_red_combined = all_red_events + all_red_from_fouls
            team_a_red = len([e for e in all_red_combined 
                            if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()])
            team_b_red = len([e for e in all_red_combined 
                            if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()])
            
            # Define disciplinary stats based on language
            if current_lang == "EN":
                stats_disciplinary = [
                    ('Fouls', team_a_fouls, team_b_fouls),
                    ('Yellow Cards', team_a_yellow, team_b_yellow),
                    ('Red Cards', team_a_red, team_b_red),
                ]
            else:  # TH
                stats_disciplinary = [
                    ('ฟาวล์', team_a_fouls, team_b_fouls),
                    ('ใบเหลือง', team_a_yellow, team_b_yellow),
                    ('ใบแดง', team_a_red, team_b_red),
                ]
            
            for stat_name, val_a, val_b in stats_disciplinary:
                ws[f'A{row}'] = stat_name
                ws[f'B{row}'] = val_a
                ws[f'C{row}'] = val_b
                ws[f'D{row}'] = f'=B{row}-C{row}'
                if val_a + val_b > 0:
                    ws[f'E{row}'] = f'=IF(B{row}+C{row}>0, B{row}/(B{row}+C{row})*100, 0)'
                else:
                    ws[f'E{row}'] = 0
                ws[f'F{row}'] = f'=IF(D{row}>0, "{team_a}", IF(D{row}<0, "{team_b}", "{equal_text}"))'
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{row}'].border = border
                    if col in ['B', 'C', 'D', 'E']:
                        ws[f'{col}{row}'].alignment = center_align
                row += 1
            row += 1
            
            # === 5. สถิติการโจมตี (Attacking Statistics) ===
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = attacking_stats_text
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row += 1
            
            attacking_types = ["ยิง", "ส่งบอล", "ข้ามบอล", "ผ่านบอล", "ส่งบอลยาว", "ส่งบอลในเขตโทษ", "เตะมุม", "ฟรีคิก"]
            all_attacking = [e for e in self.events if e.event_type in attacking_types]
            team_a_attacking = [e for e in all_attacking 
                              if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()]
            team_b_attacking = [e for e in all_attacking 
                              if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()]
            
            all_corners = [e for e in self.events if e.event_type == "เตะมุม"]
            team_a_corners = len([e for e in all_corners 
                                if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()])
            team_b_corners = len([e for e in all_corners 
                                if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()])
            
            all_free_kicks = [e for e in self.events if e.event_type == "ฟรีคิก"]
            team_a_free_kicks = len([e for e in all_free_kicks 
                                   if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()])
            team_b_free_kicks = len([e for e in all_free_kicks 
                                   if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()])
            
            all_penalties = [e for e in self.events if e.event_type == "ลูกโทษ"]
            team_a_penalties = len([e for e in all_penalties 
                                  if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()])
            team_b_penalties = len([e for e in all_penalties 
                                  if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()])
            
            # Define attacking stats based on language
            if current_lang == "EN":
                stats_attacking = [
                    ('Corner Kicks', team_a_corners, team_b_corners),
                    ('Free Kicks', team_a_free_kicks, team_b_free_kicks),
                    ('Penalties', team_a_penalties, team_b_penalties),
                    ('Total Attacking Actions', len(team_a_attacking), len(team_b_attacking)),
                ]
            else:  # TH
                stats_attacking = [
                    ('เตะมุม', team_a_corners, team_b_corners),
                    ('ฟรีคิก', team_a_free_kicks, team_b_free_kicks),
                    ('ลูกโทษ', team_a_penalties, team_b_penalties),
                    ('รวมการโจมตี', len(team_a_attacking), len(team_b_attacking)),
                ]
            
            for stat_name, val_a, val_b in stats_attacking:
                ws[f'A{row}'] = stat_name
                ws[f'B{row}'] = val_a
                ws[f'C{row}'] = val_b
                ws[f'D{row}'] = f'=B{row}-C{row}'
                if val_a + val_b > 0:
                    ws[f'E{row}'] = f'=IF(B{row}+C{row}>0, B{row}/(B{row}+C{row})*100, 0)'
                else:
                    ws[f'E{row}'] = 0
                ws[f'F{row}'] = f'=IF(D{row}>0, "{team_a}", IF(D{row}<0, "{team_b}", "{equal_text}"))'
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{row}'].border = border
                    if col in ['B', 'C', 'D', 'E']:
                        ws[f'{col}{row}'].alignment = center_align
                row += 1
            row += 1
            
            # === 6. สรุปภาพรวม (Overall Summary) ===
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = overall_summary_text
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            row += 1
            
            # Calculate total events
            total_a = len(team_a_events)
            total_b = len(team_b_events)
            
            # Count total goals from all actions (not just shots)
            all_goals_events = [e for e in self.events if e.outcome == "ประตู"]
            team_a_total_goals = len([e for e in all_goals_events 
                                    if e.team and str(e.team).strip().lower() == str(team_a).strip().lower()])
            team_b_total_goals = len([e for e in all_goals_events 
                                    if e.team and str(e.team).strip().lower() == str(team_b).strip().lower()])
            
            # Define summary stats based on language
            if current_lang == "EN":
                summary_stats = [
                    ('Total Events', total_a, total_b),
                    ('Total Goals (All Types)', team_a_total_goals, team_b_total_goals),
                    ('Goals from Shots', team_a_goals, team_b_goals),
                    ('Goal Conversion Rate (Goals/Shots)', 
                     f'=IF(B{row_total_shots}>0, B{row_goals}/B{row_total_shots}*100, 0)',
                     f'=IF(C{row_total_shots}>0, C{row_goals}/C{row_total_shots}*100, 0)'),
                ]
            else:  # TH
                summary_stats = [
                    ('จำนวนเหตุการณ์ทั้งหมด', total_a, total_b),
                    ('จำนวนประตู (รวมทุกประเภท)', team_a_total_goals, team_b_total_goals),
                    ('จำนวนประตูจากการยิง', team_a_goals, team_b_goals),
                    ('อัตราการยิงได้ประตู (ประตู/การยิง)', 
                     f'=IF(B{row_total_shots}>0, B{row_goals}/B{row_total_shots}*100, 0)',
                     f'=IF(C{row_total_shots}>0, C{row_goals}/C{row_total_shots}*100, 0)'),
                ]
            
            for stat_name, val_a, val_b in summary_stats:
                ws[f'A{row}'] = stat_name
                if isinstance(val_a, (int, float)):
                    ws[f'B{row}'] = val_a
                else:
                    ws[f'B{row}'] = val_a
                if isinstance(val_b, (int, float)):
                    ws[f'C{row}'] = val_b
                else:
                    ws[f'C{row}'] = val_b
                ws[f'D{row}'] = f'=B{row}-C{row}'
                ws[f'E{row}'] = f'=IF(B{row}+C{row}>0, B{row}/(B{row}+C{row})*100, 0)'
                ws[f'F{row}'] = f'=IF(D{row}>0, "{team_a}", IF(D{row}<0, "{team_b}", "{equal_text}"))'
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{row}'].border = border
                    ws[f'{col}{row}'].font = Font(bold=True)
                    if col in ['B', 'C', 'D', 'E']:
                        ws[f'{col}{row}'].alignment = center_align
                row += 1
            
            # Auto-adjust column widths
            from openpyxl.utils import get_column_letter
            for col in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                # Check all cells in this column
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)
                        except:
                            pass
                
                # Set width with padding (min 10, max 50)
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A3'  # Freeze after title and header row
            
            # Workbook will be saved automatically when context manager exits
        except Exception as e:
            # If openpyxl operations fail, just continue without the professional sheet
            import traceback
            print(f"Warning: Could not add professional statistics sheet: {e}")
            print(traceback.format_exc())
    
    def get_statistics(self) -> Dict:
        """Calculate statistics from tracking data"""
        stats = {
            'total_events': len(self.events),
            'events_by_type': {},
            'events_by_team': {'Team A': 0, 'Team B': 0, 'Neutral': 0},
            'events_by_half': {1: 0, 2: 0},
            'goals': {'Team A': 0, 'Team B': 0},
            'shots': {'Team A': 0, 'Team B': 0},
            'passes': {'Team A': 0, 'Team B': 0},
            'tackles': {'Team A': 0, 'Team B': 0},
            'fouls': {'Team A': 0, 'Team B': 0},
        }
        
        for event in self.events:
            # Count by type
            if event.event_type not in stats['events_by_type']:
                stats['events_by_type'][event.event_type] = 0
            stats['events_by_type'][event.event_type] += 1
            
            # Count by team
            if event.team in stats['events_by_team']:
                stats['events_by_team'][event.team] += 1
            
            # Count by half
            if event.half in stats['events_by_half']:
                stats['events_by_half'][event.half] += 1
            
            # Specific event counts
            if event.event_type == 'Goal' and event.team in stats['goals']:
                stats['goals'][event.team] += 1
            elif event.event_type == 'Shot' and event.team in stats['shots']:
                stats['shots'][event.team] += 1
            elif event.event_type == 'Pass' and event.team in stats['passes']:
                stats['passes'][event.team] += 1
            elif event.event_type == 'Tackle' and event.team in stats['tackles']:
                stats['tackles'][event.team] += 1
            elif event.event_type == 'Foul' and event.team in stats['fouls']:
                stats['fouls'][event.team] += 1
        
        return stats

